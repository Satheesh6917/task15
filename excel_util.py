import openpyxl
from datetime import datetime

class ExcelUtil:
    def __init__(self, file_path):
        self.file_path = file_path
        self.wb = openpyxl.load_workbook(file_path)
        self.sheet = self.wb.active

    def get_test_data(self):
        data = []
        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            data.append(row)
        return data

    def write_result(self, row, result):
        self.sheet.cell(row=row, column=7, value=result)
        self.sheet.cell(row=row, column=4, value=datetime.now().strftime("%Y-%m-%d"))
        self.sheet.cell(row=row, column=5, value=datetime.now().strftime("%H:%M:%S"))
        self.wb.save(self.file_path)
